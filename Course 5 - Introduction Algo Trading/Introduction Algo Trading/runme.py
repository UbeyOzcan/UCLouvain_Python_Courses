from funcs.Extractor import extract_stock_price
from funcs.Bollinger import bollinger
from funcs.RSI import rsi
from funcs.Strategy import strategy
from funcs.Graph import plot_startegy
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

pd.options.plotting.backend = "plotly"

ticker_to_trade = '^FCHI'
start_period = '2022-01-01'
end_period = '2023-10-19'
df = extract_stock_price(ticker=ticker_to_trade,
                         start=start_period,
                         end=end_period)

window_bb = 20
delta = 2
window_rsi = 6
rsi_lower_bound = 20
rsi_upper_bound = 70
sl_1 = False
sl_2 = True
sl_perc = 0.95

bb = bollinger(df_prices=df,
               window=window_bb,
               delta=delta)

rsi = rsi(df_prices=bb,
          window=window_rsi)

result_strat_1 = strategy(df_prices=rsi,
                          rsi_lower=rsi_lower_bound,
                          rsi_upper=rsi_upper_bound,
                          stop_loss=sl_1)

result_strat_2 = strategy(df_prices=rsi,
                          rsi_lower=rsi_lower_bound,
                          rsi_upper=rsi_upper_bound,
                          stop_loss=sl_2,
                          sl=sl_perc)

fig1_strat_1, fig2_strat_1, fig3_strat_1 = plot_startegy(result_strat_1, rsi_l=rsi_lower_bound, rsi_u=rsi_upper_bound)
fig1_strat_2, fig2_strat_2, fig3_strat_2 = plot_startegy(result_strat_2, rsi_l=rsi_lower_bound, rsi_u=rsi_upper_bound)

fig1_strat_1.write_image(file='plot_strat_1_bb.png', format='png')
fig2_strat_1.write_image(file='plot_strat_1_rsi.png', format='png')
fig3_strat_1.write_image(file='plot_strat_1_strat.png', format='png')

fig1_strat_2.write_image(file='plot_strat_2_bb.png', format='png')
fig2_strat_2.write_image(file='plot_strat_2_rsi.png', format='png')
fig3_strat_2.write_image(file='plot_strat_2_strat.png', format='png')


result_strat_1["df"].to_excel(f'Result of Strategy one {ticker_to_trade} for {[start_period, end_period]}.xlsx', sheet_name='Dataset')
result_strat_2["df"].to_excel(f'Result of Strategy two {ticker_to_trade} for {[start_period, end_period]}.xlsx', sheet_name='Dataset')

with pd.ExcelWriter(
        f"Result of Strategy one {ticker_to_trade} for {[start_period, end_period]}.xlsx",
        mode="a",
        engine="openpyxl",
        if_sheet_exists="replace",
) as writer:
    result_strat_1["buy"].to_excel(writer, sheet_name='buy')
    result_strat_1["sell"].to_excel(writer, sheet_name='sell')

with pd.ExcelWriter(
        f"Result of Strategy two {ticker_to_trade} for {[start_period, end_period]}.xlsx",
        mode="a",
        engine="openpyxl",
        if_sheet_exists="replace",
) as writer:
    result_strat_2["buy"].to_excel(writer, sheet_name='buy')
    result_strat_2["sell"].to_excel(writer, sheet_name='sell')

wb = load_workbook(f'Result of Strategy one {ticker_to_trade} for {[start_period, end_period]}.xlsx')
sheet1 = wb.create_sheet('graph', 0)
active = wb['graph']
active.add_image(Image('plot_strat_1_bb.png'), 'A1')
active.add_image(Image('plot_strat_1_rsi.png'), 'M1')
active.add_image(Image('plot_strat_1_strat.png'), 'A29')
summary_sheet = wb.create_sheet('Summary', 0)
summary_sheet['A1'] = 'Strategy Name : '
summary_sheet['B1'] = 'Bollinger Band and RSI'
summary_sheet['A2'] = 'Return : '
summary_sheet['B2'] = result_strat_1['return']
summary_sheet['A3'] = 'Bollinger Bands Window : '
summary_sheet['B3'] = window_bb
summary_sheet['A4'] = 'Bollinger Bands Delta : '
summary_sheet['B4'] = delta
summary_sheet['A5'] = 'RSI Window : '
summary_sheet['B5'] = window_rsi
summary_sheet['A6'] = 'RSI Lower and Upper bands : '
summary_sheet['B6'] = f'{[rsi_lower_bound, rsi_upper_bound]}'
summary_sheet['A7'] = 'Stop-Loss : '
summary_sheet['B7'] = sl_1
summary_sheet['A8'] = 'Stock traded : '
summary_sheet['B8'] = ticker_to_trade
summary_sheet['A9'] = 'Period : '
summary_sheet['B9'] = f'{[start_period, end_period]}'

wb.save(f'Result of Strategy one {ticker_to_trade} for {[start_period, end_period]}.xlsx')

wb = load_workbook(f'Result of Strategy two {ticker_to_trade} for {[start_period, end_period]}.xlsx')
sheet1 = wb.create_sheet('graph', 0)
active = wb['graph']
active.add_image(Image('plot_strat_2_bb.png'), 'A1')
active.add_image(Image('plot_strat_2_rsi.png'), 'M1')
active.add_image(Image('plot_strat_2_strat.png'), 'A29')
summary_sheet = wb.create_sheet('summary', 0)
summary_sheet['A1'] = 'Strategy Name : '
summary_sheet['B1'] = 'Bollinger Band and RSI'
summary_sheet['A2'] = 'Return : '
summary_sheet['B2'] = result_strat_2['return']
summary_sheet['A3'] = 'Bollinger Bands Window : '
summary_sheet['B3'] = window_bb
summary_sheet['A4'] = 'Bollinger Bands Delta : '
summary_sheet['B4'] = delta
summary_sheet['A5'] = 'RSI Window : '
summary_sheet['B5'] = window_rsi
summary_sheet['A6'] = 'RSI Lower and Upper bands : '
summary_sheet['B6'] = f'{[rsi_lower_bound, rsi_upper_bound]}'
summary_sheet['A7'] = 'Stop-Loss : '
summary_sheet['B7'] = sl_2
summary_sheet['A8'] = 'Stop-Loss % : '
summary_sheet['B8'] = sl_perc
summary_sheet['A9'] = 'Stock traded : '
summary_sheet['B9'] = ticker_to_trade
summary_sheet['A10'] = 'Period : '
summary_sheet['B10'] = f'{[start_period, end_period]}'
wb.save(f'Result of Strategy two {ticker_to_trade} for {[start_period, end_period]}.xlsx')


print('Your strategy has been computed !')
